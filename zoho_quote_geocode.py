#!/usr/bin/env python3

from __future__ import annotations

import argparse
import datetime as dt
import json
import logging
import os
import sys
import time
from contextlib import ExitStack
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlsplit, urlunsplit

import httpx
import shapefile

APP_NAME = "update-quote-geolocation"
APP_VERSION = os.getenv("UPDATE_QUOTE_GEOLOCATION_VERSION", "0.1.0")
ENV_FILE_ENVVAR = "ZOHO_QUOTE_GEOLOCATION_ENV_FILE"
DEFAULT_ENV_PATHS = (
    Path("/etc") / APP_NAME / "zoho_quote_geocode.env",
    Path.home() / ".config" / APP_NAME / "zoho_quote_geocode.env",
    Path.cwd() / "zoho_quote_geocode.env",
)


class ConfigError(RuntimeError):
    """Raised when required configuration is missing or invalid."""


class ZohoApiError(RuntimeError):
    """Raised when a Zoho CRM API request fails."""


class GoogleGeocodeError(RuntimeError):
    """Raised when a Google Geocoding API request fails."""


def _read_env(*names: str, default: str | None = None) -> str | None:
    for name in names:
        value = os.getenv(name)
        if value:
            return value
    return default


def _parse_env_assignment(line: str) -> tuple[str, str] | None:
    stripped = line.strip()
    if not stripped or stripped.startswith("#") or "=" not in stripped:
        return None
    key, value = stripped.split("=", 1)
    key = key.strip()
    value = value.strip()
    if value and value[0] == value[-1] and value[0] in {"'", '"'}:
        value = value[1:-1]
    if key:
        return key, value
    return None


def _load_env_file(path: Path) -> bool:
    if not path.exists():
        return False

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        parsed = _parse_env_assignment(raw_line)
        if not parsed:
            continue
        key, value = parsed
        os.environ.setdefault(key, value)
    return True


def _load_default_env_files() -> list[Path]:
    loaded: list[Path] = []
    configured_path = _read_env(ENV_FILE_ENVVAR)
    if configured_path:
        target = Path(configured_path).expanduser()
        if _load_env_file(target):
            loaded.append(target)
        return loaded

    for path in DEFAULT_ENV_PATHS:
        if _load_env_file(path):
            loaded.append(path)
    return loaded


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r", " ").replace("\n", ", ").strip()
    text = " ".join(text.split())
    return text or None


def _coerce_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _json_default(value: Any) -> Any:
    if isinstance(value, (dt.date, dt.datetime, dt.time)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, Decimal):
        return float(value)
    return str(value)


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, default=_json_default),
        encoding="utf-8",
    )


def _json_string(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=_json_default) if value is not None else ""


def _normalize_coordinate(value: float, decimal_places: int, max_length: int) -> float:
    try:
        decimal_value = Decimal(str(value))
    except InvalidOperation as exc:
        raise ConfigError(f"Invalid coordinate value: {value}") from exc

    quantizer = Decimal("1").scaleb(-decimal_places)
    rounded = decimal_value.quantize(quantizer, rounding=ROUND_HALF_UP)
    rendered = format(rounded, "f")
    if "." in rendered:
        rendered = rendered.rstrip("0").rstrip(".")

    if len(rendered) > max_length:
        raise ConfigError(
            f"Coordinate '{rendered}' exceeds the configured Zoho field length limit of {max_length}. "
            "Lower ZOHO_QUOTE_COORD_DECIMALS or increase the CRM field length."
        )

    return float(rendered)


@dataclass(slots=True)
class ZohoAuthConfig:
    api_base_url: str
    accounts_url: str
    module_api_name: str
    access_token: str | None
    refresh_token: str | None
    client_id: str | None
    client_secret: str | None
    page_size: int
    timeout_seconds: float


@dataclass(slots=True)
class QuoteFieldConfig:
    street_field: str
    city_field: str
    state_field: str
    postal_code_field: str
    country_field: str
    latitude_field: str | None
    longitude_field: str | None
    region_name_field: str | None
    region_code_field: str | None
    mrc_name_field: str | None
    muni_name_field: str | None
    arrond_name_field: str | None
    coordinate_decimal_places: int
    coordinate_max_length: int

    def requested_fields(self) -> list[str]:
        fields = [
            self.street_field,
            self.city_field,
            self.state_field,
            self.postal_code_field,
            self.country_field,
        ]
        if self.latitude_field:
            fields.append(self.latitude_field)
        if self.longitude_field:
            fields.append(self.longitude_field)
        if self.region_name_field:
            fields.append(self.region_name_field)
        if self.region_code_field:
            fields.append(self.region_code_field)
        if self.mrc_name_field:
            fields.append(self.mrc_name_field)
        if self.muni_name_field:
            fields.append(self.muni_name_field)
        if self.arrond_name_field:
            fields.append(self.arrond_name_field)
        seen: set[str] = set()
        ordered: list[str] = []
        for field_name in fields:
            if field_name not in seen:
                seen.add(field_name)
                ordered.append(field_name)
        return ordered


@dataclass(slots=True)
class QuoteAddressRecord:
    quote_id: str
    shipping_street: str | None
    shipping_city: str | None
    shipping_state: str | None
    shipping_postal_code: str | None
    shipping_country: str | None
    current_latitude: float | None
    current_longitude: float | None
    current_region_name: str | None
    current_region_code: str | None
    current_mrc_name: str | None
    current_muni_name: str | None
    current_arrond_name: str | None
    address_fields: dict[str, Any]

    @classmethod
    def from_zoho_record(cls, record: dict[str, Any], fields: QuoteFieldConfig) -> "QuoteAddressRecord":
        address_fields = {
            fields.street_field: record.get(fields.street_field),
            fields.city_field: record.get(fields.city_field),
            fields.state_field: record.get(fields.state_field),
            fields.postal_code_field: record.get(fields.postal_code_field),
            fields.country_field: record.get(fields.country_field),
        }
        if fields.latitude_field:
            address_fields[fields.latitude_field] = record.get(fields.latitude_field)
        if fields.longitude_field:
            address_fields[fields.longitude_field] = record.get(fields.longitude_field)
        if fields.region_name_field:
            address_fields[fields.region_name_field] = record.get(fields.region_name_field)
        if fields.region_code_field:
            address_fields[fields.region_code_field] = record.get(fields.region_code_field)
        if fields.mrc_name_field:
            address_fields[fields.mrc_name_field] = record.get(fields.mrc_name_field)
        if fields.muni_name_field:
            address_fields[fields.muni_name_field] = record.get(fields.muni_name_field)
        if fields.arrond_name_field:
            address_fields[fields.arrond_name_field] = record.get(fields.arrond_name_field)

        return cls(
            quote_id=str(record["id"]),
            shipping_street=_clean_text(record.get(fields.street_field)),
            shipping_city=_clean_text(record.get(fields.city_field)),
            shipping_state=_clean_text(record.get(fields.state_field)),
            shipping_postal_code=_clean_text(record.get(fields.postal_code_field)),
            shipping_country=_clean_text(record.get(fields.country_field)),
            current_latitude=_coerce_float(record.get(fields.latitude_field)) if fields.latitude_field else None,
            current_longitude=_coerce_float(record.get(fields.longitude_field)) if fields.longitude_field else None,
            current_region_name=_clean_text(record.get(fields.region_name_field)) if fields.region_name_field else None,
            current_region_code=_clean_text(record.get(fields.region_code_field)) if fields.region_code_field else None,
            current_mrc_name=_clean_text(record.get(fields.mrc_name_field)) if fields.mrc_name_field else None,
            current_muni_name=_clean_text(record.get(fields.muni_name_field)) if fields.muni_name_field else None,
            current_arrond_name=_clean_text(record.get(fields.arrond_name_field)) if fields.arrond_name_field else None,
            address_fields=address_fields,
        )

    def formatted_address(self) -> str | None:
        locality = ", ".join(part for part in [self.shipping_city, self.shipping_state] if part)
        if self.shipping_postal_code:
            locality = f"{locality} {self.shipping_postal_code}".strip() if locality else self.shipping_postal_code
        parts = [self.shipping_street, locality or None, self.shipping_country]
        rendered = ", ".join(part for part in parts if part)
        return rendered or None

    def has_coordinates(self) -> bool:
        return self.current_latitude is not None and self.current_longitude is not None

    def missing_shipping_fields(self, fields: "QuoteFieldConfig") -> list[str]:
        missing: list[str] = []
        candidates = [
            (fields.street_field, self.shipping_street),
            (fields.city_field, self.shipping_city),
            (fields.state_field, self.shipping_state),
            (fields.postal_code_field, self.shipping_postal_code),
            (fields.country_field, self.shipping_country),
        ]
        for field_name, value in candidates:
            if not value:
                missing.append(field_name)
        return missing

    def to_dict(self) -> dict[str, Any]:
        return {
            "quote_id": self.quote_id,
            "formatted_address": self.formatted_address(),
            "shipping_street": self.shipping_street,
            "shipping_city": self.shipping_city,
            "shipping_state": self.shipping_state,
            "shipping_postal_code": self.shipping_postal_code,
            "shipping_country": self.shipping_country,
            "current_latitude": self.current_latitude,
            "current_longitude": self.current_longitude,
            "current_region_name": self.current_region_name,
            "current_region_code": self.current_region_code,
            "current_mrc_name": self.current_mrc_name,
            "current_muni_name": self.current_muni_name,
            "current_arrond_name": self.current_arrond_name,
            "address_fields": self.address_fields,
        }


@dataclass(slots=True)
class RegionLookupConfig:
    source_label: str
    shape_path: Path
    region_name_attribute: str
    region_code_attribute: str | None
    mrc_name_attribute: str | None = None
    muni_name_attribute: str | None = None
    arrond_name_attribute: str | None = None


@dataclass(slots=True)
class RegionMatch:
    source_label: str
    name: str | None
    code: str | None
    mrc_name: str | None
    muni_name: str | None
    arrond_name: str | None
    attributes: dict[str, Any]


@dataclass(slots=True)
class GeocodeResult:
    latitude: float
    longitude: float
    formatted_address: str
    place_id: str | None
    location_type: str | None

    def to_dict(self) -> dict[str, Any]:
        return {
            "latitude": self.latitude,
            "longitude": self.longitude,
            "formatted_address": self.formatted_address,
            "place_id": self.place_id,
            "location_type": self.location_type,
        }


def _split_ring_points(points: list[tuple[float, float]], parts: list[int]) -> list[list[tuple[float, float]]]:
    ring_starts = list(parts) + [len(points)]
    rings: list[list[tuple[float, float]]] = []
    for index in range(len(ring_starts) - 1):
        ring = points[ring_starts[index] : ring_starts[index + 1]]
        if ring:
            rings.append(ring)
    return rings


def _ring_signed_area(ring: list[tuple[float, float]]) -> float:
    if len(ring) < 3:
        return 0.0
    area = 0.0
    for index, (x1, y1) in enumerate(ring):
        x2, y2 = ring[(index + 1) % len(ring)]
        area += (x1 * y2) - (x2 * y1)
    return area / 2.0


def _point_in_ring(x: float, y: float, ring: list[tuple[float, float]]) -> bool:
    inside = False
    j = len(ring) - 1
    for i in range(len(ring)):
        xi, yi = ring[i]
        xj, yj = ring[j]
        intersects = ((yi > y) != (yj > y)) and (
            x < (xj - xi) * (y - yi) / ((yj - yi) or 1e-12) + xi
        )
        if intersects:
            inside = not inside
        j = i
    return inside


def _group_polygon_rings(rings: list[list[tuple[float, float]]]) -> list[tuple[list[tuple[float, float]], list[list[tuple[float, float]]]]]:
    polygons: list[tuple[list[tuple[float, float]], list[list[tuple[float, float]]]]] = []
    current_outer: list[tuple[float, float]] | None = None
    current_holes: list[list[tuple[float, float]]] = []

    for ring in rings:
        signed_area = _ring_signed_area(ring)
        is_outer = signed_area < 0
        if is_outer or current_outer is None:
            if current_outer is not None:
                polygons.append((current_outer, current_holes))
            current_outer = ring
            current_holes = []
        else:
            current_holes.append(ring)

    if current_outer is not None:
        polygons.append((current_outer, current_holes))

    return polygons


class RegionShapeResolver:
    def __init__(self, config: RegionLookupConfig, logger: logging.Logger) -> None:
        self.config = config
        self.logger = logger
        self.shape_path = self._resolve_shape_path(config.shape_path)
        self.reader = shapefile.Reader(str(self.shape_path), encoding="utf-8")
        self.fields = [field[0] for field in self.reader.fields[1:]]
        self.name_index = self._field_index(config.region_name_attribute)
        self.code_index = self._field_index(config.region_code_attribute) if config.region_code_attribute else None
        self.mrc_index = self._field_index(config.mrc_name_attribute) if config.mrc_name_attribute else None
        self.muni_index = self._field_index(config.muni_name_attribute) if config.muni_name_attribute else None
        self.arrond_index = self._field_index(config.arrond_name_attribute) if config.arrond_name_attribute else None
        self.records = list(self.reader.iterShapeRecords())

    def __enter__(self) -> "RegionShapeResolver":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.reader.close()

    def lookup(self, longitude: float, latitude: float) -> RegionMatch | None:
        for shape_record in self.records:
            shape = shape_record.shape
            min_x, min_y, max_x, max_y = shape.bbox
            if longitude < min_x or longitude > max_x or latitude < min_y or latitude > max_y:
                continue

            rings = _split_ring_points(shape.points, list(shape.parts))
            polygons = _group_polygon_rings(rings)
            for outer, holes in polygons:
                if not _point_in_ring(longitude, latitude, outer):
                    continue
                if any(_point_in_ring(longitude, latitude, hole) for hole in holes):
                    continue

                record_values = list(shape_record.record)
                attributes = dict(zip(self.fields, record_values))
                name = _clean_text(record_values[self.name_index]) if self.name_index is not None else None
                code = (
                    _clean_text(record_values[self.code_index])
                    if self.code_index is not None
                    else None
                )
                mrc_name = (
                    _clean_text(record_values[self.mrc_index])
                    if self.mrc_index is not None
                    else None
                )
                muni_name = (
                    _clean_text(record_values[self.muni_index])
                    if self.muni_index is not None
                    else None
                )
                arrond_name = (
                    _clean_text(record_values[self.arrond_index])
                    if self.arrond_index is not None
                    else None
                )
                return RegionMatch(
                    source_label=self.config.source_label,
                    name=name,
                    code=code,
                    mrc_name=mrc_name,
                    muni_name=muni_name,
                    arrond_name=arrond_name,
                    attributes=attributes,
                )

        return None

    def _field_index(self, field_name: str | None) -> int | None:
        if not field_name:
            return None
        try:
            return self.fields.index(field_name)
        except ValueError as exc:
            raise ConfigError(
                f"Shapefile attribute '{field_name}' was not found in {self.shape_path.name}. "
                f"Available attributes: {', '.join(self.fields)}"
            ) from exc

    @staticmethod
    def _resolve_shape_path(path: Path) -> Path:
        candidate = path.expanduser()
        if candidate.is_dir():
            shp_files = sorted(candidate.glob("*.shp"))
            if len(shp_files) != 1:
                raise ConfigError(
                    f"Expected exactly one .shp file in directory '{candidate}', found {len(shp_files)}."
                )
            candidate = shp_files[0]
        if candidate.suffix.lower() != ".shp":
            raise ConfigError(f"Shape path must point to a .shp file or a folder containing one: {candidate}")
        if not candidate.exists():
            raise ConfigError(f"Shapefile not found: {candidate}")
        return candidate


class ZohoCrmClient:
    def __init__(self, config: ZohoAuthConfig, field_config: QuoteFieldConfig, logger: logging.Logger) -> None:
        self.config = config
        self.field_config = field_config
        self.logger = logger
        timeout = httpx.Timeout(self.config.timeout_seconds, connect=min(20.0, self.config.timeout_seconds))
        self._client = httpx.Client(timeout=timeout)
        self._access_token: str | None = self.config.access_token

    def close(self) -> None:
        self._client.close()

    def __enter__(self) -> "ZohoCrmClient":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def fetch_quotes_with_shipping_addresses(self, max_records: int | None = None) -> list[QuoteAddressRecord]:
        fields = ",".join(self.field_config.requested_fields())
        records: list[QuoteAddressRecord] = []
        page = 1
        next_page_token: str | None = None
        page_size = max(1, min(self.config.page_size, 200))

        while True:
            params: dict[str, Any] = {
                "fields": fields,
                "per_page": page_size,
            }
            if next_page_token:
                params["page_token"] = next_page_token
            else:
                params["page"] = page

            payload = self._request(
                "GET",
                f"{self.config.api_base_url}/{self.config.module_api_name}",
                params=params,
            )

            data = payload.get("data") or []
            if not data:
                break

            for raw_record in data:
                records.append(QuoteAddressRecord.from_zoho_record(raw_record, self.field_config))
                if max_records is not None and len(records) >= max_records:
                    return records

            info = payload.get("info") or {}
            next_page_token = info.get("next_page_token") or payload.get("next_page_token")
            more_records = bool(info.get("more_records")) or bool(next_page_token)
            if not more_records:
                break
            if not next_page_token:
                page += 1

        return records

    def update_quote_coordinates(self, quote_id: str, latitude: float, longitude: float) -> dict[str, Any]:
        if not self.field_config.latitude_field or not self.field_config.longitude_field:
            raise ConfigError(
                "Latitude and longitude field API names are required for sync mode. "
                "Set ZOHO_QUOTE_LATITUDE_FIELD and ZOHO_QUOTE_LONGITUDE_FIELD, or pass "
                "--latitude-field and --longitude-field."
            )

        safe_latitude = _normalize_coordinate(
            latitude,
            decimal_places=self.field_config.coordinate_decimal_places,
            max_length=self.field_config.coordinate_max_length,
        )
        safe_longitude = _normalize_coordinate(
            longitude,
            decimal_places=self.field_config.coordinate_decimal_places,
            max_length=self.field_config.coordinate_max_length,
        )

        payload = {
            "data": [
                {
                    self.field_config.latitude_field: safe_latitude,
                    self.field_config.longitude_field: safe_longitude,
                }
            ]
        }

        return self._request(
            "PUT",
            f"{self.config.api_base_url}/{self.config.module_api_name}/{quote_id}",
            json=payload,
        )

    def update_quote_fields(self, quote_id: str, values: dict[str, Any]) -> dict[str, Any]:
        if not values:
            raise ConfigError("No values were provided for the Zoho CRM update.")

        payload = {
            "data": [values]
        }
        return self._request(
            "PUT",
            f"{self.config.api_base_url}/{self.config.module_api_name}/{quote_id}",
            json=payload,
        )

    def _request(self, method: str, url: str, **kwargs: Any) -> dict[str, Any]:
        headers = dict(kwargs.pop("headers", {}))
        headers["Authorization"] = f"Zoho-oauthtoken {self._get_access_token()}"
        response = self._client.request(method, url, headers=headers, **kwargs)
        if response.status_code >= 400:
            raise ZohoApiError(self._format_error(response))
        if response.status_code == 204 or not response.content:
            return {}
        return response.json()

    def _get_access_token(self) -> str:
        if self._access_token:
            return self._access_token

        if not self.config.refresh_token or not self.config.client_id or not self.config.client_secret:
            raise ConfigError(
                "Missing Zoho OAuth credentials. Set ZOHO_CRM_ACCESS_TOKEN, or set "
                "ZOHO_CRM_REFRESH_TOKEN, ZOHO_CRM_CLIENT_ID, and ZOHO_CRM_CLIENT_SECRET."
            )

        response = self._client.post(
            self.config.accounts_url,
            data={
                "grant_type": "refresh_token",
                "refresh_token": self.config.refresh_token,
                "client_id": self.config.client_id,
                "client_secret": self.config.client_secret,
            },
        )
        if response.status_code >= 400:
            raise ZohoApiError(self._format_error(response, prefix="Zoho OAuth refresh failed"))

        payload = response.json()
        access_token = payload.get("access_token")
        if not access_token:
            raise ZohoApiError(f"Zoho OAuth refresh did not return an access token: {payload}")

        api_domain = payload.get("api_domain")
        if api_domain:
            self.config.api_base_url = self._replace_base_domain(self.config.api_base_url, str(api_domain))

        self._access_token = str(access_token)
        return self._access_token

    @staticmethod
    def _format_error(response: httpx.Response, prefix: str = "Zoho CRM API request failed") -> str:
        body: Any
        try:
            body = response.json()
        except ValueError:
            body = response.text
        return f"{prefix} with status {response.status_code}: {body}"

    @staticmethod
    def _replace_base_domain(api_base_url: str, api_domain: str) -> str:
        current = urlsplit(api_base_url)
        refreshed = urlsplit(api_domain)
        return urlunsplit((refreshed.scheme, refreshed.netloc, current.path, current.query, current.fragment))


class GoogleGeocoder:
    def __init__(
        self,
        api_key: str,
        logger: logging.Logger,
        timeout_seconds: float = 30.0,
        max_retries: int = 3,
        retry_delay_seconds: float = 1.0,
    ) -> None:
        self.api_key = api_key
        self.logger = logger
        self.max_retries = max(1, max_retries)
        self.retry_delay_seconds = max(0.0, retry_delay_seconds)
        timeout = httpx.Timeout(timeout_seconds, connect=min(20.0, timeout_seconds))
        self._client = httpx.Client(timeout=timeout)

    def close(self) -> None:
        self._client.close()

    def __enter__(self) -> "GoogleGeocoder":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def geocode(self, address: str) -> GeocodeResult | None:
        for attempt in range(1, self.max_retries + 1):
            response = self._client.get(
                "https://maps.googleapis.com/maps/api/geocode/json",
                params={
                    "address": address,
                    "key": self.api_key,
                },
            )
            if response.status_code >= 400:
                if attempt < self.max_retries and response.status_code >= 500:
                    time.sleep(self.retry_delay_seconds)
                    continue
                raise GoogleGeocodeError(
                    f"Google Geocoding request failed with status {response.status_code}: {response.text}"
                )

            payload = response.json()
            status = payload.get("status")
            if status == "OK":
                result = payload["results"][0]
                location = (result.get("geometry") or {}).get("location") or {}
                return GeocodeResult(
                    latitude=float(location["lat"]),
                    longitude=float(location["lng"]),
                    formatted_address=str(result.get("formatted_address") or address),
                    place_id=result.get("place_id"),
                    location_type=(result.get("geometry") or {}).get("location_type"),
                )

            if status == "ZERO_RESULTS":
                return None

            if status == "UNKNOWN_ERROR" and attempt < self.max_retries:
                time.sleep(self.retry_delay_seconds)
                continue

            error_message = payload.get("error_message")
            detail = f"{status}: {error_message}" if error_message else str(status)
            raise GoogleGeocodeError(f"Google Geocoding failed for '{address}': {detail}")

        raise GoogleGeocodeError(f"Google Geocoding failed for '{address}' after retries")


def fetch_quote_shipping_addresses(
    zoho_client: ZohoCrmClient,
    max_records: int | None = None,
) -> list[QuoteAddressRecord]:
    return zoho_client.fetch_quotes_with_shipping_addresses(max_records=max_records)


def _build_record_item(record: QuoteAddressRecord, fields: QuoteFieldConfig) -> dict[str, Any]:
    item = record.to_dict()
    missing_fields = record.missing_shipping_fields(fields)
    item["missing_shipping_fields"] = missing_fields
    item["has_missing_shipping_fields"] = bool(missing_fields)
    missing_coordinate_fields: list[str] = []
    if fields.latitude_field and record.current_latitude is None:
        missing_coordinate_fields.append(fields.latitude_field)
    if fields.longitude_field and record.current_longitude is None:
        missing_coordinate_fields.append(fields.longitude_field)
    item["missing_coordinate_fields"] = missing_coordinate_fields
    item["has_missing_coordinate_fields"] = bool(missing_coordinate_fields)
    return item


def _admin_target_specs(
    record: QuoteAddressRecord,
    fields: QuoteFieldConfig,
    match: RegionMatch | None = None,
) -> list[tuple[str, str | None, str | None]]:
    specs: list[tuple[str, str | None, str | None]] = []
    if fields.region_name_field:
        specs.append((fields.region_name_field, record.current_region_name, match.name if match else None))
    if fields.region_code_field:
        specs.append((fields.region_code_field, record.current_region_code, match.code if match else None))
    if fields.mrc_name_field:
        specs.append((fields.mrc_name_field, record.current_mrc_name, match.mrc_name if match else None))
    if fields.muni_name_field:
        specs.append((fields.muni_name_field, record.current_muni_name, match.muni_name if match else None))
    if fields.arrond_name_field:
        specs.append((fields.arrond_name_field, record.current_arrond_name, match.arrond_name if match else None))
    return specs


def _remaining_admin_fields(
    record: QuoteAddressRecord,
    fields: QuoteFieldConfig,
    update_values: dict[str, Any] | None = None,
) -> list[str]:
    effective_updates = update_values or {}
    missing_fields: list[str] = []
    for field_name, current_value, _ in _admin_target_specs(record, fields):
        final_value = effective_updates.get(field_name, current_value)
        if not _clean_text(final_value):
            missing_fields.append(field_name)
    return missing_fields


def _merge_boundary_matches(
    boundary_match: RegionMatch | None,
    arrond_match: RegionMatch | None,
) -> RegionMatch | None:
    if boundary_match is None and arrond_match is None:
        return None

    if boundary_match is not None:
        merged_attributes = dict(boundary_match.attributes)
        if arrond_match is not None:
            merged_attributes["arrondissement"] = arrond_match.attributes
        return RegionMatch(
            source_label=boundary_match.source_label,
            name=boundary_match.name,
            code=boundary_match.code,
            mrc_name=boundary_match.mrc_name,
            muni_name=boundary_match.muni_name,
            arrond_name=arrond_match.arrond_name or arrond_match.name if arrond_match else None,
            attributes=merged_attributes,
        )

    return RegionMatch(
        source_label=arrond_match.source_label,
        name=None,
        code=None,
        mrc_name=None,
        muni_name=None,
        arrond_name=arrond_match.arrond_name or arrond_match.name,
        attributes={"arrondissement": arrond_match.attributes},
    )


def _should_include_in_failure_report(item: dict[str, Any]) -> bool:
    if item.get("missing_shipping_fields"):
        return True
    return item.get("status") in {
        "skipped_missing_address",
        "skipped_missing_coordinates",
        "no_geocode_result",
        "geocode_error",
        "region_lookup_error",
        "no_region_match",
        "no_admin_update_values",
        "updated_partial",
        "update_error",
    }


def _write_failure_report(path: Path, payload: dict[str, Any], logger: logging.Logger) -> tuple[Path, int]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:  # pragma: no cover - dependency should be installed by package
        raise ConfigError(
            "openpyxl is required to create the Excel failure report. "
            "Install dependencies again or upgrade the package."
        ) from exc

    issue_rows = [item for item in payload.get("items", []) if _should_include_in_failure_report(item)]
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["metric", "value"])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
    for key, value in (payload.get("summary") or {}).items():
        summary_sheet.append([key, value])
    summary_sheet.append(["issue_rows", len(issue_rows)])
    summary_sheet.freeze_panes = "A2"

    issues_sheet = workbook.create_sheet("issues")
    headers = [
        "quote_id",
        "status",
        "status_reason",
        "google_status",
        "missing_shipping_fields",
        "missing_coordinate_fields",
        "missing_admin_fields",
        "formatted_address",
        "shipping_street",
        "shipping_city",
        "shipping_state",
        "shipping_postal_code",
        "shipping_country",
        "current_latitude",
        "current_longitude",
        "current_region_name",
        "current_region_code",
        "current_mrc_name",
        "current_muni_name",
        "current_arrond_name",
        "geocoded_latitude",
        "geocoded_longitude",
        "geocoded_formatted_address",
        "geocode_place_id",
        "geocode_location_type",
        "resolved_region_name",
        "resolved_region_code",
        "resolved_mrc_name",
        "resolved_muni_name",
        "resolved_arrond_name",
        "region_match_source",
        "arrond_match_source",
        "error",
        "raw_address_fields",
        "update_response",
    ]
    issues_sheet.append(headers)
    for cell in issues_sheet[1]:
        cell.font = Font(bold=True)

    for item in issue_rows:
        geocode = item.get("geocode") or {}
        issues_sheet.append(
            [
                item.get("quote_id", ""),
                item.get("status", ""),
                item.get("status_reason", ""),
                item.get("google_status", ""),
                ", ".join(item.get("missing_shipping_fields") or []),
                ", ".join(item.get("missing_coordinate_fields") or []),
                ", ".join(item.get("missing_admin_fields") or []),
                item.get("formatted_address", ""),
                item.get("shipping_street", ""),
                item.get("shipping_city", ""),
                item.get("shipping_state", ""),
                item.get("shipping_postal_code", ""),
                item.get("shipping_country", ""),
                item.get("current_latitude", ""),
                item.get("current_longitude", ""),
                item.get("current_region_name", ""),
                item.get("current_region_code", ""),
                item.get("current_mrc_name", ""),
                item.get("current_muni_name", ""),
                item.get("current_arrond_name", ""),
                geocode.get("latitude", ""),
                geocode.get("longitude", ""),
                geocode.get("formatted_address", ""),
                geocode.get("place_id", ""),
                geocode.get("location_type", ""),
                item.get("resolved_region_name", ""),
                item.get("resolved_region_code", ""),
                item.get("resolved_mrc_name", ""),
                item.get("resolved_muni_name", ""),
                item.get("resolved_arrond_name", ""),
                item.get("region_match_source", ""),
                item.get("arrond_match_source", ""),
                item.get("error", ""),
                _json_string(item.get("address_fields")),
                _json_string(item.get("update_response")),
            ]
        )

    for sheet in (summary_sheet, issues_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 48)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    logger.info("Wrote Excel failure report to %s with %s issue rows", path, len(issue_rows))
    return path, len(issue_rows)


def _should_include_in_google_error_report(item: dict[str, Any]) -> bool:
    return item.get("status") in {"no_geocode_result", "geocode_error"}


def _write_google_error_report(path: Path, payload: dict[str, Any], logger: logging.Logger) -> tuple[Path, int]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:  # pragma: no cover - dependency should be installed by package
        raise ConfigError(
            "openpyxl is required to create the Excel Google error report. "
            "Install dependencies again or upgrade the package."
        ) from exc

    issue_rows = [item for item in payload.get("items", []) if _should_include_in_google_error_report(item)]
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["metric", "value"])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
    for key, value in (payload.get("summary") or {}).items():
        if key in {"fetched", "no_geocode_result", "geocode_errors"}:
            summary_sheet.append([key, value])
    summary_sheet.append(["google_issue_rows", len(issue_rows)])
    summary_sheet.freeze_panes = "A2"

    issues_sheet = workbook.create_sheet("google_errors")
    headers = [
        "quote_id",
        "status",
        "status_reason",
        "google_status",
        "formatted_address",
        "shipping_street",
        "shipping_city",
        "shipping_state",
        "shipping_postal_code",
        "shipping_country",
        "missing_shipping_fields",
        "error",
        "raw_address_fields",
    ]
    issues_sheet.append(headers)
    for cell in issues_sheet[1]:
        cell.font = Font(bold=True)

    for item in issue_rows:
        issues_sheet.append(
            [
                item.get("quote_id", ""),
                item.get("status", ""),
                item.get("status_reason", ""),
                item.get("google_status", ""),
                item.get("formatted_address", ""),
                item.get("shipping_street", ""),
                item.get("shipping_city", ""),
                item.get("shipping_state", ""),
                item.get("shipping_postal_code", ""),
                item.get("shipping_country", ""),
                ", ".join(item.get("missing_shipping_fields") or []),
                item.get("error", ""),
                _json_string(item.get("address_fields")),
            ]
        )

    for sheet in (summary_sheet, issues_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 48)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    logger.info("Wrote Google geocode error report to %s with %s issue rows", path, len(issue_rows))
    return path, len(issue_rows)


def _load_json_payload(path: Path) -> dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ConfigError(f"JSON input file was not found: {path}") from exc
    except json.JSONDecodeError as exc:
        raise ConfigError(f"JSON input file is not valid JSON: {path}") from exc


def _ordered_quote_ids(*payloads: dict[str, Any] | None) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for payload in payloads:
        for item in payload.get("items", []) if payload else []:
            quote_id = str(item.get("quote_id") or "")
            if quote_id and quote_id not in seen:
                seen.add(quote_id)
                ordered.append(quote_id)
    return ordered


def _compact_statuses(items: list[dict[str, Any]]) -> str:
    parts: list[str] = []
    for item in items:
        status = item.get("status")
        quote_id = item.get("quote_id")
        if status and quote_id:
            parts.append(f"{quote_id} ({status})")
    return ", ".join(parts)


def _build_run_report(
    *,
    sync_payload: dict[str, Any] | None,
    region_payload: dict[str, Any] | None,
    sync_input_path: Path | None = None,
    region_input_path: Path | None = None,
) -> dict[str, Any]:
    sync_items = {
        str(item.get("quote_id")): item
        for item in (sync_payload or {}).get("items", [])
        if item.get("quote_id")
    }
    region_items = {
        str(item.get("quote_id")): item
        for item in (region_payload or {}).get("items", [])
        if item.get("quote_id")
    }

    summary: dict[str, Any] = {}
    if sync_payload:
        for key, value in (sync_payload.get("summary") or {}).items():
            summary[f"sync_{key}"] = value
    if region_payload:
        for key, value in (region_payload.get("summary") or {}).items():
            summary[f"region_{key}"] = value

    quote_rows: list[dict[str, Any]] = []
    issue_rows: list[dict[str, Any]] = []
    quote_ids = _ordered_quote_ids(sync_payload, region_payload)

    for quote_id in quote_ids:
        sync_item = sync_items.get(quote_id) or {}
        region_item = region_items.get(quote_id) or {}
        base_item = region_item or sync_item

        row = {
            "quote_id": quote_id,
            "formatted_address": base_item.get("formatted_address") or sync_item.get("formatted_address"),
            "shipping_street": base_item.get("shipping_street") or sync_item.get("shipping_street"),
            "shipping_city": base_item.get("shipping_city") or sync_item.get("shipping_city"),
            "shipping_state": base_item.get("shipping_state") or sync_item.get("shipping_state"),
            "shipping_postal_code": base_item.get("shipping_postal_code") or sync_item.get("shipping_postal_code"),
            "shipping_country": base_item.get("shipping_country") or sync_item.get("shipping_country"),
            "sync_status": sync_item.get("status"),
            "sync_status_reason": sync_item.get("status_reason"),
            "google_status": sync_item.get("google_status"),
            "geocode_request_address": sync_item.get("geocode_request_address"),
            "start_latitude": sync_item.get("current_latitude"),
            "start_longitude": sync_item.get("current_longitude"),
            "geocoded_latitude": (sync_item.get("geocode") or {}).get("latitude"),
            "geocoded_longitude": (sync_item.get("geocode") or {}).get("longitude"),
            "coordinate_update_values": sync_item.get("coordinate_update_values"),
            "sync_error": sync_item.get("error"),
            "region_status": region_item.get("status"),
            "region_status_reason": region_item.get("status_reason"),
            "region_match_source": region_item.get("region_match_source"),
            "arrond_match_source": region_item.get("arrond_match_source"),
            "region_start_latitude": region_item.get("current_latitude"),
            "region_start_longitude": region_item.get("current_longitude"),
            "current_region_name": region_item.get("current_region_name"),
            "current_region_code": region_item.get("current_region_code"),
            "current_mrc_name": region_item.get("current_mrc_name"),
            "current_muni_name": region_item.get("current_muni_name"),
            "current_arrond_name": region_item.get("current_arrond_name"),
            "resolved_region_name": region_item.get("resolved_region_name"),
            "resolved_region_code": region_item.get("resolved_region_code"),
            "resolved_mrc_name": region_item.get("resolved_mrc_name"),
            "resolved_muni_name": region_item.get("resolved_muni_name"),
            "resolved_arrond_name": region_item.get("resolved_arrond_name"),
            "admin_update_values": region_item.get("admin_update_values"),
            "region_error": region_item.get("error"),
            "sync_started_missing_shipping_fields": sync_item.get("missing_shipping_fields") or region_item.get("missing_shipping_fields") or [],
            "sync_started_missing_coordinate_fields": sync_item.get("missing_coordinate_fields") or [],
            "remaining_admin_fields_after_region": region_item.get("missing_admin_fields") or [],
        }

        if row["sync_status"] in {"skipped_missing_address", "no_geocode_result", "geocode_error", "update_error"}:
            row["overall_outcome"] = "sync_issue"
        elif row["region_status"] in {"region_lookup_error", "no_region_match", "update_error"}:
            row["overall_outcome"] = "region_issue"
        elif row["region_status"] in {"updated_partial", "no_admin_update_values"}:
            row["overall_outcome"] = "partial_boundary"
        elif row["sync_status"] == "skipped_existing_coordinates":
            row["overall_outcome"] = "used_existing_coordinates"
        else:
            row["overall_outcome"] = "ok"

        quote_rows.append(row)
        if row["overall_outcome"] != "ok":
            issue_rows.append(row)

    summary_lines: list[str] = []
    if sync_payload:
        sync_summary = sync_payload.get("summary") or {}
        summary_lines.append(
            "Geocode sync processed "
            f"{sync_summary.get('fetched', 0)} quotes: "
            f"{sync_summary.get('updated', 0)} updated, "
            f"{sync_summary.get('skipped_existing_coordinates', 0)} already had coordinates, "
            f"{sync_summary.get('skipped_missing_address', 0)} missing addresses, "
            f"{sync_summary.get('no_geocode_result', 0)} Google zero-results, "
            f"{sync_summary.get('geocode_errors', 0)} Google API errors, and "
            f"{sync_summary.get('update_errors', 0)} Zoho coordinate update errors."
        )
        sync_blockers = [
            item
            for item in (sync_payload.get("items") or [])
            if item.get("status") in {"skipped_missing_address", "no_geocode_result", "geocode_error", "update_error"}
        ]
        if sync_blockers:
            summary_lines.append("Quotes blocked during geocode sync: " + _compact_statuses(sync_blockers))
        elif sync_summary.get("updated", 0) or sync_summary.get("skipped_existing_coordinates", 0):
            summary_lines.append("Google geocoding returned a usable result for every address sent in this run.")

    if region_payload:
        region_summary = region_payload.get("summary") or {}
        summary_lines.append(
            "Boundary sync processed "
            f"{region_summary.get('fetched', 0)} quotes: "
            f"{region_summary.get('updated', 0)} full boundary updates, "
            f"{region_summary.get('updated_partial', 0)} partial updates, "
            f"{region_summary.get('no_admin_update_values', 0)} no-op updates, "
            f"{region_summary.get('skipped_missing_coordinates', 0)} skipped for missing coordinates, "
            f"{region_summary.get('region_lookup_errors', 0)} lookup errors, and "
            f"{region_summary.get('update_errors', 0)} Zoho boundary update errors."
        )
        partial_items = [
            item
            for item in (region_payload.get("items") or [])
            if item.get("status") == "updated_partial"
        ]
        if partial_items and all((item.get("missing_admin_fields") or []) == ["Arrondissement"] for item in partial_items):
            summary_lines.append(
                f"All {len(partial_items)} partial boundary updates were missing only Arrondissement. "
                "Region, MRC, and Muni were resolved and written successfully."
            )
        region_blockers = [
            item
            for item in (region_payload.get("items") or [])
            if item.get("status") in {"skipped_missing_coordinates", "region_lookup_error", "no_region_match", "update_error", "no_admin_update_values"}
        ]
        if region_blockers:
            summary_lines.append("Quotes requiring review during boundary sync: " + _compact_statuses(region_blockers))

    if sync_payload and region_payload:
        sync_ids = {str(item.get("quote_id")) for item in (sync_payload.get("items") or []) if item.get("quote_id")}
        region_ids = {str(item.get("quote_id")) for item in (region_payload.get("items") or []) if item.get("quote_id")}
        if sync_ids != region_ids:
            summary_lines.append(
                "The sync and region-sync inputs do not cover the exact same quote set. "
                "Compare quote IDs before drawing conclusions across both steps."
            )

    return {
        "meta": {
            "app_name": APP_NAME,
            "app_version": APP_VERSION,
            "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
            "sync_input_path": str(sync_input_path) if sync_input_path else None,
            "region_input_path": str(region_input_path) if region_input_path else None,
        },
        "summary": summary,
        "summary_lines": summary_lines,
        "quotes": quote_rows,
        "issues": issue_rows,
        "source_payloads": {
            "sync": sync_payload,
            "region_sync": region_payload,
        },
    }


def _write_run_report(path: Path, payload: dict[str, Any], logger: logging.Logger) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:  # pragma: no cover
        raise ConfigError(
            "openpyxl is required to create the consolidated Excel run report. "
            "Install dependencies again or upgrade the package."
        ) from exc

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["section", "key", "value"])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)

    for key, value in (payload.get("meta") or {}).items():
        summary_sheet.append(["meta", key, value])

    for index, line in enumerate(payload.get("summary_lines") or [], start=1):
        summary_sheet.append(["narrative", f"line_{index}", line])

    for key, value in (payload.get("summary") or {}).items():
        summary_sheet.append(["metric", key, value])

    quotes_sheet = workbook.create_sheet("quotes")
    quote_rows = payload.get("quotes") or []
    quote_headers = [
        "quote_id",
        "overall_outcome",
        "formatted_address",
        "shipping_street",
        "shipping_city",
        "shipping_state",
        "shipping_postal_code",
        "shipping_country",
        "sync_status",
        "sync_status_reason",
        "google_status",
        "geocode_request_address",
        "start_latitude",
        "start_longitude",
        "geocoded_latitude",
        "geocoded_longitude",
        "coordinate_update_values",
        "sync_error",
        "region_status",
        "region_status_reason",
        "region_match_source",
        "arrond_match_source",
        "region_start_latitude",
        "region_start_longitude",
        "current_region_name",
        "current_region_code",
        "current_mrc_name",
        "current_muni_name",
        "current_arrond_name",
        "resolved_region_name",
        "resolved_region_code",
        "resolved_mrc_name",
        "resolved_muni_name",
        "resolved_arrond_name",
        "admin_update_values",
        "region_error",
            "sync_started_missing_shipping_fields",
            "sync_started_missing_coordinate_fields",
            "remaining_admin_fields_after_region",
        ]
    quotes_sheet.append(quote_headers)
    for cell in quotes_sheet[1]:
        cell.font = Font(bold=True)
    for row in quote_rows:
        quotes_sheet.append(
            [
                row.get("quote_id", ""),
                row.get("overall_outcome", ""),
                row.get("formatted_address", ""),
                row.get("shipping_street", ""),
                row.get("shipping_city", ""),
                row.get("shipping_state", ""),
                row.get("shipping_postal_code", ""),
                row.get("shipping_country", ""),
                row.get("sync_status", ""),
                row.get("sync_status_reason", ""),
                row.get("google_status", ""),
                row.get("geocode_request_address", ""),
                row.get("start_latitude", ""),
                row.get("start_longitude", ""),
                row.get("geocoded_latitude", ""),
                row.get("geocoded_longitude", ""),
                _json_string(row.get("coordinate_update_values")),
                row.get("sync_error", ""),
                row.get("region_status", ""),
                row.get("region_status_reason", ""),
                row.get("region_match_source", ""),
                row.get("arrond_match_source", ""),
                row.get("region_start_latitude", ""),
                row.get("region_start_longitude", ""),
                row.get("current_region_name", ""),
                row.get("current_region_code", ""),
                row.get("current_mrc_name", ""),
                row.get("current_muni_name", ""),
                row.get("current_arrond_name", ""),
                row.get("resolved_region_name", ""),
                row.get("resolved_region_code", ""),
                row.get("resolved_mrc_name", ""),
                row.get("resolved_muni_name", ""),
                row.get("resolved_arrond_name", ""),
                _json_string(row.get("admin_update_values")),
                row.get("region_error", ""),
                ", ".join(row.get("sync_started_missing_shipping_fields") or []),
                ", ".join(row.get("sync_started_missing_coordinate_fields") or []),
                ", ".join(row.get("remaining_admin_fields_after_region") or []),
            ]
        )

    issues_sheet = workbook.create_sheet("issues")
    issues_sheet.append(quote_headers)
    for cell in issues_sheet[1]:
        cell.font = Font(bold=True)
    for row in payload.get("issues") or []:
        issues_sheet.append(
            [
                row.get("quote_id", ""),
                row.get("overall_outcome", ""),
                row.get("formatted_address", ""),
                row.get("shipping_street", ""),
                row.get("shipping_city", ""),
                row.get("shipping_state", ""),
                row.get("shipping_postal_code", ""),
                row.get("shipping_country", ""),
                row.get("sync_status", ""),
                row.get("sync_status_reason", ""),
                row.get("google_status", ""),
                row.get("geocode_request_address", ""),
                row.get("start_latitude", ""),
                row.get("start_longitude", ""),
                row.get("geocoded_latitude", ""),
                row.get("geocoded_longitude", ""),
                _json_string(row.get("coordinate_update_values")),
                row.get("sync_error", ""),
                row.get("region_status", ""),
                row.get("region_status_reason", ""),
                row.get("region_match_source", ""),
                row.get("arrond_match_source", ""),
                row.get("region_start_latitude", ""),
                row.get("region_start_longitude", ""),
                row.get("current_region_name", ""),
                row.get("current_region_code", ""),
                row.get("current_mrc_name", ""),
                row.get("current_muni_name", ""),
                row.get("current_arrond_name", ""),
                row.get("resolved_region_name", ""),
                row.get("resolved_region_code", ""),
                row.get("resolved_mrc_name", ""),
                row.get("resolved_muni_name", ""),
                row.get("resolved_arrond_name", ""),
                _json_string(row.get("admin_update_values")),
                row.get("region_error", ""),
                ", ".join(row.get("sync_started_missing_shipping_fields") or []),
                ", ".join(row.get("sync_started_missing_coordinate_fields") or []),
                ", ".join(row.get("remaining_admin_fields_after_region") or []),
            ]
        )

    raw_sheet = workbook.create_sheet("raw_json")
    raw_sheet.append(["source", "line_number", "content"])
    for cell in raw_sheet[1]:
        cell.font = Font(bold=True)
    for source_name, source_payload in (payload.get("source_payloads") or {}).items():
        if not source_payload:
            continue
        for line_number, line in enumerate(
            json.dumps(source_payload, indent=2, ensure_ascii=False, default=_json_default).splitlines(),
            start=1,
        ):
            raw_sheet.append([source_name, line_number, line])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 60)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    logger.info("Wrote consolidated run report to %s", path)
    return path


def geocode_quote_records(
    records: Iterable[QuoteAddressRecord],
    geocoder: GoogleGeocoder,
    *,
    field_config: QuoteFieldConfig,
    skip_existing: bool = True,
) -> list[dict[str, Any]]:
    report: list[dict[str, Any]] = []

    for record in records:
        item = _build_record_item(record, field_config)
        address = record.formatted_address()

        if not address:
            item["status"] = "skipped_missing_address"
            report.append(item)
            continue

        if skip_existing and record.has_coordinates():
            item["status"] = "skipped_existing_coordinates"
            report.append(item)
            continue

        geocode = geocoder.geocode(address)
        if geocode is None:
            item["status"] = "no_geocode_result"
            report.append(item)
            continue

        item["status"] = "geocoded"
        item["geocode"] = geocode.to_dict()
        report.append(item)

    return report


def sync_quote_coordinates(
    zoho_client: ZohoCrmClient,
    geocoder: GoogleGeocoder,
    *,
    max_records: int | None = None,
    skip_existing: bool = True,
    dry_run: bool = False,
) -> dict[str, Any]:
    logger = zoho_client.logger
    records = fetch_quote_shipping_addresses(zoho_client, max_records=max_records)
    summary = {
        "fetched": len(records),
        "updated": 0,
        "dry_run": 0,
        "skipped_missing_address": 0,
        "skipped_existing_coordinates": 0,
        "no_geocode_result": 0,
        "geocode_errors": 0,
        "update_errors": 0,
    }
    items: list[dict[str, Any]] = []

    for record in records:
        item = _build_record_item(record, zoho_client.field_config)
        address = record.formatted_address()
        item["geocode_request_address"] = address

        if not address:
            item["status"] = "skipped_missing_address"
            item["status_reason"] = "No formatted address could be built from the configured shipping fields."
            summary["skipped_missing_address"] += 1
            logger.warning(
                "Quote %s skipped in geocode sync: missing address fields: %s",
                record.quote_id,
                ", ".join(item.get("missing_shipping_fields") or []),
            )
            items.append(item)
            continue

        if skip_existing and record.has_coordinates():
            item["status"] = "skipped_existing_coordinates"
            item["status_reason"] = "Quote already has both latitude and longitude."
            summary["skipped_existing_coordinates"] += 1
            logger.info(
                "Quote %s skipped in geocode sync: latitude/longitude already populated.",
                record.quote_id,
            )
            items.append(item)
            continue

        try:
            geocode = geocoder.geocode(address)
        except GoogleGeocodeError as exc:
            item["status"] = "geocode_error"
            item["status_reason"] = "Google Geocoding API returned an error."
            item["google_status"] = "ERROR"
            item["error"] = str(exc)
            summary["geocode_errors"] += 1
            logger.warning(
                "Quote %s geocode error for address '%s': %s",
                record.quote_id,
                address,
                exc,
            )
            items.append(item)
            continue

        if geocode is None:
            item["status"] = "no_geocode_result"
            item["status_reason"] = "Google Geocoding returned ZERO_RESULTS for the formatted address."
            item["google_status"] = "ZERO_RESULTS"
            summary["no_geocode_result"] += 1
            logger.warning(
                "Quote %s geocode returned ZERO_RESULTS for address '%s'.",
                record.quote_id,
                address,
            )
            items.append(item)
            continue

        item["geocode"] = geocode.to_dict()
        item["google_status"] = "OK"
        item["coordinate_update_values"] = {
            "latitude": geocode.latitude,
            "longitude": geocode.longitude,
        }

        if dry_run:
            item["status"] = "dry_run"
            item["status_reason"] = "Geocoding succeeded, but CRM update was skipped because dry-run mode is enabled."
            summary["dry_run"] += 1
            logger.info(
                "Quote %s geocoded successfully in dry-run mode: lat=%s long=%s",
                record.quote_id,
                geocode.latitude,
                geocode.longitude,
            )
            items.append(item)
            continue

        try:
            update_response = zoho_client.update_quote_coordinates(
                record.quote_id,
                geocode.latitude,
                geocode.longitude,
            )
        except ZohoApiError as exc:
            item["status"] = "update_error"
            item["status_reason"] = "Zoho CRM rejected the latitude/longitude update."
            item["error"] = str(exc)
            summary["update_errors"] += 1
            logger.warning(
                "Quote %s geocoded but Zoho update failed: %s",
                record.quote_id,
                exc,
            )
            items.append(item)
            continue

        item["status"] = "updated"
        item["status_reason"] = "Latitude and longitude were updated in Zoho CRM."
        item["update_response"] = update_response
        summary["updated"] += 1
        logger.info(
            "Quote %s updated with latitude=%s longitude=%s",
            record.quote_id,
            geocode.latitude,
            geocode.longitude,
        )
        items.append(item)

    return {
        "summary": summary,
        "items": items,
    }


def sync_quote_regions(
    zoho_client: ZohoCrmClient,
    resolvers: list[RegionShapeResolver],
    *,
    arrond_resolver: RegionShapeResolver | None = None,
    max_records: int | None = None,
    update_existing: bool = False,
) -> dict[str, Any]:
    logger = zoho_client.logger
    if not zoho_client.field_config.latitude_field or not zoho_client.field_config.longitude_field:
        raise ConfigError(
            "Latitude and longitude field API names are required for region sync. "
            "Set ZOHO_QUOTE_LATITUDE_FIELD and ZOHO_QUOTE_LONGITUDE_FIELD."
        )
    if not any(
        [
            zoho_client.field_config.region_name_field,
            zoho_client.field_config.region_code_field,
            zoho_client.field_config.mrc_name_field,
            zoho_client.field_config.muni_name_field,
            zoho_client.field_config.arrond_name_field,
        ]
    ):
        raise ConfigError(
            "At least one admin-boundary target field is required for region sync. "
            "Set one or more of ZOHO_QUOTE_REGION_NAME_FIELD, ZOHO_QUOTE_REGION_CODE_FIELD, "
            "ZOHO_QUOTE_MRC_NAME_FIELD, ZOHO_QUOTE_MUNI_NAME_FIELD, or ZOHO_QUOTE_ARRON_NAME_FIELD."
        )
    if zoho_client.field_config.arrond_name_field and arrond_resolver is None:
        raise ConfigError(
            "An arrondissement target field was configured, but no arrondissement shapefile was provided. "
            "Set ZOHO_ARRON_SHAPE_PATH or pass --arron-shape-path."
        )

    records = fetch_quote_shipping_addresses(zoho_client, max_records=max_records)
    summary = {
        "fetched": len(records),
        "updated": 0,
        "updated_partial": 0,
        "skipped_missing_coordinates": 0,
        "skipped_existing_admin_fields": 0,
        "no_region_match": 0,
        "no_admin_update_values": 0,
        "region_lookup_errors": 0,
        "update_errors": 0,
        "matched_by_arrondissement": 0,
        "matched_by_muni": 0,
        "matched_by_mrc": 0,
        "matched_by_region": 0,
    }
    items: list[dict[str, Any]] = []

    for record in records:
        item = _build_record_item(record, zoho_client.field_config)
        item["missing_admin_fields"] = _remaining_admin_fields(record, zoho_client.field_config)

        if record.current_latitude is None or record.current_longitude is None:
            item["status"] = "skipped_missing_coordinates"
            item["status_reason"] = "Quote is missing latitude and/or longitude, so no polygon lookup was attempted."
            summary["skipped_missing_coordinates"] += 1
            logger.warning(
                "Quote %s skipped in region sync: missing latitude/longitude.",
                record.quote_id,
            )
            items.append(item)
            continue

        configured_admin_targets = _admin_target_specs(record, zoho_client.field_config)
        if not update_existing and configured_admin_targets and all(current_value for _, current_value, _ in configured_admin_targets):
            item["status"] = "skipped_existing_admin_fields"
            item["status_reason"] = "All requested admin-boundary fields already have values."
            item["missing_admin_fields"] = []
            summary["skipped_existing_admin_fields"] += 1
            logger.info(
                "Quote %s skipped in region sync: all requested admin fields already populated.",
                record.quote_id,
            )
            items.append(item)
            continue

        try:
            boundary_match = None
            for resolver in resolvers:
                boundary_match = resolver.lookup(record.current_longitude, record.current_latitude)
                if boundary_match is not None:
                    break
        except Exception as exc:  # pragma: no cover - unexpected shape parsing/runtime failure
            item["status"] = "region_lookup_error"
            item["status_reason"] = "The shapefile lookup failed while resolving the quote point."
            item["error"] = str(exc)
            summary["region_lookup_errors"] += 1
            logger.warning("Quote %s region lookup failed: %s", record.quote_id, exc)
            items.append(item)
            continue

        try:
            arrond_match = (
                arrond_resolver.lookup(record.current_longitude, record.current_latitude)
                if arrond_resolver is not None
                else None
            )
        except Exception as exc:  # pragma: no cover - unexpected shape parsing/runtime failure
            item["status"] = "region_lookup_error"
            item["status_reason"] = "The arrondissement shapefile lookup failed while resolving the quote point."
            item["error"] = str(exc)
            summary["region_lookup_errors"] += 1
            logger.warning("Quote %s arrondissement lookup failed: %s", record.quote_id, exc)
            items.append(item)
            continue

        if boundary_match is None:
            summary["no_region_match"] += 1
        else:
            item["region_match_source"] = boundary_match.source_label
            item["resolved_region_name"] = boundary_match.name
            item["resolved_region_code"] = boundary_match.code
            item["resolved_mrc_name"] = boundary_match.mrc_name
            item["resolved_muni_name"] = boundary_match.muni_name
            if boundary_match.source_label == "municipality":
                summary["matched_by_muni"] += 1
            elif boundary_match.source_label == "mrc":
                summary["matched_by_mrc"] += 1
            else:
                summary["matched_by_region"] += 1

        if arrond_match is not None:
            item["arrond_match_source"] = arrond_match.source_label
            item["resolved_arrond_name"] = arrond_match.arrond_name or arrond_match.name
            summary["matched_by_arrondissement"] += 1

        match = _merge_boundary_matches(boundary_match, arrond_match)
        if match is None:
            item["status"] = "no_region_match"
            item["status_reason"] = "No configured shapefile polygon contained the quote coordinates."
            item["missing_admin_fields"] = _remaining_admin_fields(record, zoho_client.field_config)
            logger.warning(
                "Quote %s had no boundary match for latitude=%s longitude=%s.",
                record.quote_id,
                record.current_latitude,
                record.current_longitude,
            )
            items.append(item)
            continue

        item["resolved_region_attributes"] = match.attributes

        update_values: dict[str, Any] = {}
        for field_name, current_value, resolved_value in _admin_target_specs(record, zoho_client.field_config, match):
            if not resolved_value:
                continue
            if update_existing or not current_value:
                update_values[field_name] = resolved_value

        item["missing_admin_fields"] = _remaining_admin_fields(record, zoho_client.field_config, update_values)
        item["admin_update_values"] = update_values

        if not update_values:
            item["status"] = "no_admin_update_values"
            item["status_reason"] = "Boundary lookup succeeded, but none of the resolved values needed to be written to Zoho."
            summary["no_admin_update_values"] += 1
            logger.info(
                "Quote %s boundary lookup resolved values, but no Zoho fields needed updating.",
                record.quote_id,
            )
            items.append(item)
            continue

        try:
            update_response = zoho_client.update_quote_fields(record.quote_id, update_values)
        except ZohoApiError as exc:
            item["status"] = "update_error"
            item["status_reason"] = "Zoho CRM rejected the boundary field update."
            item["error"] = str(exc)
            summary["update_errors"] += 1
            logger.warning("Quote %s boundary update failed: %s", record.quote_id, exc)
            items.append(item)
            continue

        item["status"] = "updated_partial" if item["missing_admin_fields"] else "updated"
        item["status_reason"] = (
            "Boundary fields were updated, but one or more requested fields are still unresolved."
            if item["missing_admin_fields"]
            else "All requested boundary fields were updated in Zoho CRM."
        )
        item["update_response"] = update_response
        if item["missing_admin_fields"]:
            summary["updated_partial"] += 1
            logger.info(
                "Quote %s boundary fields updated partially. Missing after update: %s",
                record.quote_id,
                ", ".join(item["missing_admin_fields"]),
            )
        else:
            summary["updated"] += 1
            logger.info("Quote %s boundary fields updated successfully.", record.quote_id)
        items.append(item)

    return {
        "summary": summary,
        "items": items,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Fetch Zoho CRM quotes, geocode shipping addresses, and enrich quotes with boundary fields."
    )
    parser.add_argument("--version", action="version", version=f"%(prog)s {APP_VERSION}")
    subparsers = parser.add_subparsers(dest="command", required=True)

    fetch_parser = subparsers.add_parser("fetch", help="Fetch quotes and export their shipping address fields.")
    sync_parser = subparsers.add_parser("sync", help="Fetch quotes, geocode shipping addresses, and update CRM.")
    region_parser = subparsers.add_parser(
        "region-sync",
        help="Use quote latitude/longitude to resolve arrondissement, municipality, MRC, and region polygons and update quote fields.",
    )
    report_parser = subparsers.add_parser(
        "report",
        help="Combine sync and region-sync JSON outputs into one readable Excel report.",
    )

    for current_parser in (fetch_parser, sync_parser, region_parser):
        current_parser.add_argument(
            "--api-base-url",
            default=_read_env("ZOHO_CRM_API_BASE_URL", default="https://www.zohoapis.com/crm/v7"),
            help="Zoho CRM API base URL, including the CRM version path.",
        )
        current_parser.add_argument(
            "--accounts-url",
            default=_read_env(
                "ZOHO_CRM_ACCOUNTS_URL",
                "ZOHO_WORKDRIVE_ACCOUNTS_URL",
                default="https://accounts.zoho.com/oauth/v2/token",
            ),
            help="Zoho Accounts OAuth token URL for your data center.",
        )
        current_parser.add_argument(
            "--module",
            default=_read_env("ZOHO_CRM_MODULE", default="Quotes"),
            help="Zoho CRM module API name. Default: Quotes",
        )
        current_parser.add_argument(
            "--street-field",
            default=_read_env("ZOHO_QUOTE_SHIPPING_STREET_FIELD", default="Shipping_Street"),
            help="Shipping street field API name.",
        )
        current_parser.add_argument(
            "--city-field",
            default=_read_env("ZOHO_QUOTE_SHIPPING_CITY_FIELD", default="Shipping_City"),
            help="Shipping city field API name.",
        )
        current_parser.add_argument(
            "--state-field",
            default=_read_env("ZOHO_QUOTE_SHIPPING_STATE_FIELD", default="Shipping_State"),
            help="Shipping state/province field API name.",
        )
        current_parser.add_argument(
            "--postal-field",
            default=_read_env("ZOHO_QUOTE_SHIPPING_POSTAL_CODE_FIELD", default="Shipping_Code"),
            help="Shipping postal code field API name.",
        )
        current_parser.add_argument(
            "--country-field",
            default=_read_env("ZOHO_QUOTE_SHIPPING_COUNTRY_FIELD", default="Shipping_Country"),
            help="Shipping country field API name.",
        )
        current_parser.add_argument(
            "--coordinate-decimals",
            type=int,
            default=int(_read_env("ZOHO_QUOTE_COORD_DECIMALS", default="9") or "9"),
            help="Maximum decimal places sent to Zoho for latitude/longitude fields.",
        )
        current_parser.add_argument(
            "--coordinate-max-length",
            type=int,
            default=int(_read_env("ZOHO_QUOTE_COORD_MAX_LENGTH", default="16") or "16"),
            help="Maximum total character length sent to Zoho for latitude/longitude fields.",
        )
        current_parser.add_argument(
            "--page-size",
            type=int,
            default=int(_read_env("ZOHO_CRM_PAGE_SIZE", default="200") or "200"),
            help="Zoho page size, max 200.",
        )
        current_parser.add_argument(
            "--max-records",
            type=int,
            default=None,
            help="Optional cap on the number of quotes to process.",
        )
        current_parser.add_argument(
            "--timeout",
            type=float,
            default=float(_read_env("ZOHO_CRM_TIMEOUT_SECONDS", default="30") or "30"),
            help="HTTP timeout in seconds.",
        )
        current_parser.add_argument(
            "--output",
            type=Path,
            default=None,
            help="Optional JSON output file.",
        )
        current_parser.add_argument(
            "--log-level",
            default=os.getenv("LOG_LEVEL", "INFO"),
            choices=["DEBUG", "INFO", "WARNING", "ERROR"],
            help="Log level.",
        )

    sync_parser.add_argument(
        "--google-api-key",
        default=_read_env("GOOGLE_MAPS_API_KEY", "GOOGLE_GEOCODING_API_KEY"),
        help="Google Maps Geocoding API key.",
    )
    sync_parser.add_argument(
        "--update-existing",
        action="store_true",
        help="Update quotes even if both latitude and longitude are already populated.",
    )
    sync_parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Geocode records without updating Zoho CRM.",
    )
    sync_parser.add_argument(
        "--retry-delay",
        type=float,
        default=float(_read_env("GOOGLE_GEOCODE_RETRY_DELAY_SECONDS", default="1") or "1"),
        help="Delay between Google retry attempts in seconds.",
    )
    sync_parser.add_argument(
        "--max-retries",
        type=int,
        default=int(_read_env("GOOGLE_GEOCODE_MAX_RETRIES", default="3") or "3"),
        help="Maximum Google Geocoding retry attempts.",
    )
    sync_parser.add_argument(
        "--failure-report",
        type=Path,
        default=Path(_read_env("ZOHO_QUOTE_FAILURE_REPORT_PATH", default="quote-geolocation-failures.xlsx") or "quote-geolocation-failures.xlsx"),
        help="Excel report path for quotes with missing shipping fields or failed updates.",
    )
    sync_parser.add_argument(
        "--google-error-report",
        type=Path,
        default=Path(_read_env("ZOHO_GOOGLE_ERROR_REPORT_PATH", default="quote-google-geocode-errors.xlsx") or "quote-google-geocode-errors.xlsx"),
        help="Excel report path for quotes where Google geocoding returned ZERO_RESULTS or an API error.",
    )

    for region_target_parser in (sync_parser, region_parser):
        region_target_parser.add_argument(
            "--latitude-field",
            default=_read_env("ZOHO_QUOTE_LATITUDE_FIELD"),
            help="Latitude field API name in the quote module.",
        )
        region_target_parser.add_argument(
            "--longitude-field",
            default=_read_env("ZOHO_QUOTE_LONGITUDE_FIELD"),
            help="Longitude field API name in the quote module.",
        )

    region_parser.add_argument(
        "--region-name-field",
        default=_read_env("ZOHO_QUOTE_REGION_NAME_FIELD"),
        help="Zoho quote field API name to update with the matched region name.",
    )
    region_parser.add_argument(
        "--region-code-field",
        default=_read_env("ZOHO_QUOTE_REGION_CODE_FIELD"),
        help="Zoho quote field API name to update with the matched region code.",
    )
    region_parser.add_argument(
        "--mrc-name-field",
        default=_read_env("ZOHO_QUOTE_MRC_NAME_FIELD"),
        help="Zoho quote field API name to update with the matched MRC name.",
    )
    region_parser.add_argument(
        "--muni-name-field",
        default=_read_env("ZOHO_QUOTE_MUNI_NAME_FIELD"),
        help="Zoho quote field API name to update with the matched municipality name.",
    )
    region_parser.add_argument(
        "--arron-name-field",
        default=_read_env("ZOHO_QUOTE_ARRON_NAME_FIELD"),
        help="Zoho quote field API name to update with the matched arrondissement name.",
    )
    region_parser.add_argument(
        "--arron-shape-path",
        type=Path,
        default=Path(_read_env("ZOHO_ARRON_SHAPE_PATH")) if _read_env("ZOHO_ARRON_SHAPE_PATH") else None,
        help="Optional arrondissement shapefile (.shp) path for borough-level results such as Montreal arrondissements.",
    )
    region_parser.add_argument(
        "--arron-name-attribute",
        default=_read_env("ZOHO_ARRON_NAME_ATTRIBUTE", default="ARS_NM_ARR"),
        help="Arrondissement shapefile attribute that contains the arrondissement name.",
    )
    region_parser.add_argument(
        "--muni-shape-path",
        type=Path,
        default=Path(_read_env("ZOHO_MUNI_SHAPE_PATH")) if _read_env("ZOHO_MUNI_SHAPE_PATH") else None,
        help="Primary municipality shapefile (.shp) path.",
    )
    region_parser.add_argument(
        "--muni-name-attribute",
        default=_read_env("ZOHO_MUNI_NAME_ATTRIBUTE", default="MUS_NM_MUN"),
        help="Municipality shapefile attribute that contains the municipality name.",
    )
    region_parser.add_argument(
        "--muni-mrc-attribute",
        default=_read_env("ZOHO_MUNI_MRC_ATTRIBUTE", default="MUS_NM_MRC"),
        help="Municipality shapefile attribute that contains the MRC name.",
    )
    region_parser.add_argument(
        "--muni-region-attribute",
        default=_read_env("ZOHO_MUNI_REGION_ATTRIBUTE", default="MUS_NM_REG"),
        help="Municipality shapefile attribute that contains the region name.",
    )
    region_parser.add_argument(
        "--muni-region-code-attribute",
        default=_read_env("ZOHO_MUNI_REGION_CODE_ATTRIBUTE", default="MUS_CO_REG"),
        help="Municipality shapefile attribute that contains the region code.",
    )
    region_parser.add_argument(
        "--mrc-shape-path",
        type=Path,
        default=Path(_read_env("ZOHO_MRC_SHAPE_PATH", "ZOHO_REGION_FALLBACK_SHAPE_PATH"))
        if _read_env("ZOHO_MRC_SHAPE_PATH", "ZOHO_REGION_FALLBACK_SHAPE_PATH")
        else None,
        help="Fallback MRC shapefile (.shp) path.",
    )
    region_parser.add_argument(
        "--mrc-name-attribute",
        default=_read_env("ZOHO_MRC_NAME_ATTRIBUTE", default="MRS_NM_MRC"),
        help="MRC shapefile attribute that contains the MRC name.",
    )
    region_parser.add_argument(
        "--mrc-region-attribute",
        default=_read_env("ZOHO_MRC_REGION_ATTRIBUTE", "ZOHO_REGION_FALLBACK_NAME_ATTRIBUTE", default="MRS_NM_REG"),
        help="MRC shapefile attribute that contains the region name.",
    )
    region_parser.add_argument(
        "--mrc-region-code-attribute",
        default=_read_env("ZOHO_MRC_REGION_CODE_ATTRIBUTE", "ZOHO_REGION_FALLBACK_CODE_ATTRIBUTE", default="MRS_CO_REG"),
        help="MRC shapefile attribute that contains the region code.",
    )
    region_parser.add_argument(
        "--shape-path",
        type=Path,
        default=Path(_read_env("ZOHO_REGION_SHAPE_PATH")) if _read_env("ZOHO_REGION_SHAPE_PATH") else None,
        help="Final region fallback shapefile (.shp) path.",
    )
    region_parser.add_argument(
        "--shape-name-attribute",
        default=_read_env("ZOHO_REGION_NAME_ATTRIBUTE", default="RES_NM_REG"),
        help="Region shapefile attribute that contains the region name.",
    )
    region_parser.add_argument(
        "--shape-code-attribute",
        default=_read_env("ZOHO_REGION_CODE_ATTRIBUTE", default="RES_CO_REG"),
        help="Region shapefile attribute that contains the region code.",
    )
    region_parser.add_argument(
        "--fallback-shape-path",
        type=Path,
        default=Path(_read_env("ZOHO_REGION_FALLBACK_SHAPE_PATH")) if _read_env("ZOHO_REGION_FALLBACK_SHAPE_PATH") else None,
        help="Deprecated alias for --mrc-shape-path.",
    )
    region_parser.add_argument(
        "--fallback-shape-name-attribute",
        default=_read_env("ZOHO_REGION_FALLBACK_NAME_ATTRIBUTE", default=""),
        help="Deprecated alias for --mrc-region-attribute.",
    )
    region_parser.add_argument(
        "--fallback-shape-code-attribute",
        default=_read_env("ZOHO_REGION_FALLBACK_CODE_ATTRIBUTE", default=""),
        help="Deprecated alias for --mrc-region-code-attribute.",
    )
    region_parser.add_argument(
        "--failure-report",
        type=Path,
        default=Path(_read_env("ZOHO_REGION_FAILURE_REPORT_PATH", default="quote-region-failures.xlsx") or "quote-region-failures.xlsx"),
        help="Excel report path for quotes with missing coordinates, partial boundary matches, or failed region updates.",
    )
    region_parser.add_argument(
        "--update-existing-region",
        action="store_true",
        help="Update Region, MRC, Muni, and Arrondissement fields even when the quote already has values.",
    )

    report_parser.add_argument(
        "--sync-input",
        type=Path,
        default=None,
        help="Path to a JSON file produced by the sync command.",
    )
    report_parser.add_argument(
        "--region-input",
        type=Path,
        default=None,
        help="Path to a JSON file produced by the region-sync command.",
    )
    report_parser.add_argument(
        "--report-output",
        type=Path,
        default=Path("quote-run-report.xlsx"),
        help="Path to the consolidated Excel report file.",
    )
    report_parser.add_argument(
        "--json-output",
        type=Path,
        default=None,
        help="Optional path to write the merged report payload as JSON.",
    )
    report_parser.add_argument(
        "--log-level",
        default=os.getenv("LOG_LEVEL", "INFO"),
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Log level.",
    )

    return parser


def _build_configs(args: argparse.Namespace) -> tuple[ZohoAuthConfig, QuoteFieldConfig]:
    zoho_config = ZohoAuthConfig(
        api_base_url=str(args.api_base_url).rstrip("/"),
        accounts_url=str(args.accounts_url).rstrip("/"),
        module_api_name=args.module,
        access_token=_read_env("ZOHO_CRM_ACCESS_TOKEN", "ZOHO_WORKDRIVE_ACCESS_TOKEN"),
        refresh_token=_read_env("ZOHO_CRM_REFRESH_TOKEN", "ZOHO_WORKDRIVE_REFRESH_TOKEN"),
        client_id=_read_env("ZOHO_CRM_CLIENT_ID", "ZOHO_WORKDRIVE_CLIENT_ID"),
        client_secret=_read_env("ZOHO_CRM_CLIENT_SECRET", "ZOHO_WORKDRIVE_CLIENT_SECRET"),
        page_size=args.page_size,
        timeout_seconds=args.timeout,
    )
    field_config = QuoteFieldConfig(
        street_field=args.street_field,
        city_field=args.city_field,
        state_field=args.state_field,
        postal_code_field=args.postal_field,
        country_field=args.country_field,
        latitude_field=getattr(args, "latitude_field", None),
        longitude_field=getattr(args, "longitude_field", None),
        region_name_field=getattr(args, "region_name_field", None),
        region_code_field=getattr(args, "region_code_field", None),
        mrc_name_field=getattr(args, "mrc_name_field", None),
        muni_name_field=getattr(args, "muni_name_field", None),
        arrond_name_field=getattr(args, "arron_name_field", None),
        coordinate_decimal_places=args.coordinate_decimals,
        coordinate_max_length=args.coordinate_max_length,
    )
    return zoho_config, field_config


def _build_arrond_lookup_config(args: argparse.Namespace) -> RegionLookupConfig | None:
    arron_shape_path = getattr(args, "arron_shape_path", None)
    if not arron_shape_path:
        return None
    return RegionLookupConfig(
        source_label="arrondissement",
        shape_path=Path(arron_shape_path),
        region_name_attribute=args.arron_name_attribute,
        region_code_attribute=None,
        arrond_name_attribute=args.arron_name_attribute,
    )


def _build_region_lookup_configs(args: argparse.Namespace) -> list[RegionLookupConfig]:
    configs: list[RegionLookupConfig] = []

    muni_shape_path = getattr(args, "muni_shape_path", None)
    if muni_shape_path:
        configs.append(
            RegionLookupConfig(
                source_label="municipality",
                shape_path=Path(muni_shape_path),
                region_name_attribute=args.muni_region_attribute,
                region_code_attribute=args.muni_region_code_attribute or None,
                mrc_name_attribute=args.muni_mrc_attribute or None,
                muni_name_attribute=args.muni_name_attribute or None,
            )
        )

    mrc_shape_path = getattr(args, "mrc_shape_path", None) or getattr(args, "fallback_shape_path", None)
    if mrc_shape_path:
        mrc_region_attribute = getattr(args, "mrc_region_attribute", "") or getattr(args, "fallback_shape_name_attribute", "")
        mrc_region_code_attribute = getattr(args, "mrc_region_code_attribute", "") or getattr(args, "fallback_shape_code_attribute", "")
        configs.append(
            RegionLookupConfig(
                source_label="mrc",
                shape_path=Path(mrc_shape_path),
                region_name_attribute=mrc_region_attribute or "MRS_NM_REG",
                region_code_attribute=mrc_region_code_attribute or None,
                mrc_name_attribute=args.mrc_name_attribute or None,
            )
        )

    region_shape_path = getattr(args, "shape_path", None)
    if region_shape_path:
        configs.append(
            RegionLookupConfig(
                source_label="region",
                shape_path=Path(region_shape_path),
                region_name_attribute=args.shape_name_attribute,
                region_code_attribute=args.shape_code_attribute or None,
            )
        )

    if not configs:
        raise ConfigError(
            "At least one shapefile path is required for region sync. "
            "Set ZOHO_MUNI_SHAPE_PATH, ZOHO_MRC_SHAPE_PATH, or ZOHO_REGION_SHAPE_PATH."
        )

    return configs


def _configure_logging(level_name: str) -> logging.Logger:
    logging.basicConfig(
        level=getattr(logging, level_name.upper(), logging.INFO),
        format="%(levelname)s %(message)s",
    )
    return logging.getLogger("zoho-quote-geocode")


def main(argv: list[str] | None = None) -> int:
    _load_default_env_files()
    parser = build_parser()
    args = parser.parse_args(argv)
    logger = _configure_logging(args.log_level)

    if args.command == "report":
        if not args.sync_input and not args.region_input:
            logger.error("At least one input is required for report mode. Pass --sync-input and/or --region-input.")
            return 1

        try:
            sync_payload = _load_json_payload(args.sync_input) if args.sync_input else None
            region_payload = _load_json_payload(args.region_input) if args.region_input else None
            payload = _build_run_report(
                sync_payload=sync_payload,
                region_payload=region_payload,
                sync_input_path=args.sync_input,
                region_input_path=args.region_input,
            )
            report_path = _write_run_report(args.report_output, payload, logger)
            payload["report_output_path"] = str(report_path)
            if args.json_output:
                _write_json(args.json_output, payload)
                logger.info("Wrote merged report JSON to %s", args.json_output)
            return 0
        except ConfigError as exc:
            logger.error(str(exc))
            return 1

    zoho_config, field_config = _build_configs(args)

    try:
        with ZohoCrmClient(zoho_config, field_config, logger) as zoho_client:
            if args.command == "fetch":
                records = fetch_quote_shipping_addresses(zoho_client, max_records=args.max_records)
                payload = {
                    "count": len(records),
                    "quotes": [record.to_dict() for record in records],
                }
            elif args.command == "sync":
                if not args.google_api_key:
                    raise ConfigError(
                        "Google API key is required for sync mode. Set GOOGLE_MAPS_API_KEY or pass --google-api-key."
                    )

                with GoogleGeocoder(
                    api_key=args.google_api_key,
                    logger=logger,
                    timeout_seconds=args.timeout,
                    max_retries=args.max_retries,
                    retry_delay_seconds=args.retry_delay,
                ) as geocoder:
                    payload = sync_quote_coordinates(
                        zoho_client,
                        geocoder,
                        max_records=args.max_records,
                        skip_existing=not args.update_existing,
                        dry_run=args.dry_run,
                    )
            else:
                region_configs = _build_region_lookup_configs(args)
                arrond_config = _build_arrond_lookup_config(args)
                with ExitStack() as stack:
                    resolvers = [stack.enter_context(RegionShapeResolver(config, logger)) for config in region_configs]
                    arrond_resolver = stack.enter_context(RegionShapeResolver(arrond_config, logger)) if arrond_config else None
                    payload = sync_quote_regions(
                        zoho_client,
                        resolvers,
                        arrond_resolver=arrond_resolver,
                        max_records=args.max_records,
                        update_existing=args.update_existing_region,
                    )

    except (ConfigError, ZohoApiError, GoogleGeocodeError, shapefile.ShapefileException) as exc:
        logger.error(str(exc))
        return 1

    payload["meta"] = {
        "app_name": APP_NAME,
        "app_version": APP_VERSION,
        "command": args.command,
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "module": args.module,
        "max_records": args.max_records,
    }

    if args.command in {"sync", "region-sync"} and args.failure_report:
        try:
            report_path, issue_count = _write_failure_report(args.failure_report, payload, logger)
        except ConfigError as exc:
            logger.error(str(exc))
            return 1
        payload["failure_report_path"] = str(report_path)
        payload["failure_report_issue_count"] = issue_count

    if args.command == "sync" and args.google_error_report:
        try:
            google_report_path, google_issue_count = _write_google_error_report(args.google_error_report, payload, logger)
        except ConfigError as exc:
            logger.error(str(exc))
            return 1
        payload["google_error_report_path"] = str(google_report_path)
        payload["google_error_report_issue_count"] = google_issue_count

    if args.output:
        _write_json(args.output, payload)
        logger.info("Wrote JSON output to %s", args.output)
    else:
        json.dump(payload, sys.stdout, indent=2, ensure_ascii=False, default=_json_default)
        sys.stdout.write("\n")

    if args.command in {"sync", "region-sync"}:
        summary = payload["summary"]
        if args.command == "sync":
            logger.info(
                "Fetched=%s Updated=%s DryRun=%s MissingAddress=%s ExistingCoordinates=%s NoGeocode=%s GeocodeErrors=%s UpdateErrors=%s",
                summary["fetched"],
                summary["updated"],
                summary["dry_run"],
                summary["skipped_missing_address"],
                summary["skipped_existing_coordinates"],
                summary["no_geocode_result"],
                summary["geocode_errors"],
                summary["update_errors"],
            )
            if summary["geocode_errors"] or summary["update_errors"]:
                return 1
        else:
            logger.info(
                "Fetched=%s Updated=%s UpdatedPartial=%s MissingCoordinates=%s ExistingAdmin=%s NoRegionMatch=%s NoAdminUpdates=%s RegionLookupErrors=%s UpdateErrors=%s MatchedByArrondissement=%s MatchedByMuni=%s MatchedByMrc=%s MatchedByRegion=%s",
                summary["fetched"],
                summary["updated"],
                summary["updated_partial"],
                summary["skipped_missing_coordinates"],
                summary["skipped_existing_admin_fields"],
                summary["no_region_match"],
                summary["no_admin_update_values"],
                summary["region_lookup_errors"],
                summary["update_errors"],
                summary["matched_by_arrondissement"],
                summary["matched_by_muni"],
                summary["matched_by_mrc"],
                summary["matched_by_region"],
            )
            if summary["region_lookup_errors"] or summary["update_errors"]:
                return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
